from datetime import datetime
import pandas as pd
import openpyxl as xl
import csv
import asposecells

csvFile1 = "output2All.csv"
now = str(datetime.now()).replace(' ','').replace('-','_').replace(':','_')
xlFile = 'output_Liion_'+now +'_.xlsx'
# inputfile = 'outFile1.csv'
with open('outputAll.json', encoding='utf-8') as inputfile:
    df = pd.read_json(inputfile)


# df.sort_index()
df['LC KwH'] = pd.to_numeric(df['LC KwH'])

new_df = df.sort_values(by='LC KwH', ignore_index=True)

new_df.to_csv(csvFile1, encoding='utf-8', index=False)

print(new_df.info)

wb_tmp = xl.Workbook()
ws_tmp = wb_tmp.activeconvertjsonAll2Csv.py

with open(csvFile1) as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws_tmp.append(row)

wb_tmp.save(xlFile)
# openfile as spreadsheet
wb = xl.load_workbook(xlFile)
# set worksheet
ws = wb.active
# wb.save(xlFile)

print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))